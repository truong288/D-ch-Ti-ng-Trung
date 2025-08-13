import os   #CHẠY OK
import json
import hashlib
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (Application, CommandHandler, MessageHandler,
                          ContextTypes, filters, CallbackQueryHandler)
from deep_translator import GoogleTranslator
from pypinyin import pinyin, Style
import openpyxl
from openpyxl.utils import get_column_letter
from telegram.ext import MessageHandler, filters
from stay_alive import keep_alive


keep_alive()

# ==== File cấu hình ====
ADMIN_FILE = "admins.json"
DB_FILE = "translation_db.json"
TEMP_FILE = "temp_callback_data.json"


# ==== Quản lý Admin ====
def load_admins():
    try:
        with open(ADMIN_FILE, "r", encoding="utf-8") as f:
            return set(json.load(f))
    except:
        save_admins(set())
        return set()


def save_admins(admin_set):
    with open(ADMIN_FILE, "w", encoding="utf-8") as f:
        json.dump(list(admin_set), f, ensure_ascii=False, indent=2)


admins = load_admins()


def is_admin(user_id):
    return str(user_id) in admins


# ==== Quản lý Database ====
class TranslationDatabase:

    def __init__(self):
        self.data = {"history": [], "saved_phrases": {}}
        self.load_db()

    def load_db(self):
        try:
            with open(DB_FILE, "r", encoding="utf-8") as f:
                self.data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self.save_db()

    def save_db(self):
        with open(DB_FILE, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def save_phrase(self, user_id, phrase, translation):
        if str(user_id) not in self.data["saved_phrases"]:
            self.data["saved_phrases"][str(user_id)] = {}
        self.data["saved_phrases"][str(user_id)][phrase] = translation
        self.save_db()

    def delete_phrase(self, user_id, phrase):
        if str(user_id) in self.data["saved_phrases"]:
            if phrase in self.data["saved_phrases"][str(user_id)]:
                del self.data["saved_phrases"][str(user_id)][phrase]
                self.save_db()
                return True
        return False

    def search_phrases(self, user_id, keyword):
        if str(user_id) not in self.data["saved_phrases"]:
            return {}
        return {
            k: v
            for k, v in self.data["saved_phrases"][str(user_id)].items()
            if keyword.lower() in k.lower() or keyword.lower() in v.lower()
        }

    def add_history(self, user_id, original, translated, direction):
        self.data["history"].append({
            "user_id": str(user_id),
            "original": original,
            "translated": translated,
            "direction": direction,
            "timestamp": datetime.now().isoformat()
        })
        self.save_db()

    def get_user_history(self, user_id, limit=5):
        user_history = [
            item for item in self.data["history"]
            if item["user_id"] == str(user_id)
        ]
        return sorted(user_history, key=lambda x: x["timestamp"],
                      reverse=True)[:limit]


db = TranslationDatabase()

# ==== Lưu trữ callback_data tạm bền vững ====


def load_temp_data():
    try:
        with open(TEMP_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {}


def save_temp_data(data):
    with open(TEMP_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


temp_data = load_temp_data()


# ==== Hàm hỗ trợ tạo ID ngắn từ text ====
def short_id(text):
    return hashlib.md5(text.encode()).hexdigest()[:10]


# ==== Dịch ngôn ngữ & Phiên âm ====
async def translate_chunk(text):
    try:
        if any('\u4e00' <= char <= '\u9fff' for char in text):
            translated = GoogleTranslator(source='zh-CN',
                                          target='vi').translate(text)
            pinyin_text = " ".join(
                [item[0] for item in pinyin(text, style=Style.TONE)])
            return {
                "translation": translated,
                "direction": "zh-vi",
                "pinyin": pinyin_text
            }

        vietnamese_chars = set('aăâbcdđeêghiklmnoôơpqrstuưvxy')
        if any(char in vietnamese_chars for char in text.lower()):
            translated = GoogleTranslator(source='vi',
                                          target='zh-CN').translate(text)
            pinyin_text = " ".join(
                [item[0] for item in pinyin(translated, style=Style.TONE)])
            return {
                "translation": translated,
                "direction": "vi-zh",
                "pinyin": pinyin_text
            }
    except Exception as e:
        print(f"Lỗi khi dịch: {e}")
    return None


async def detect_and_translate(text, user_id=None):
    result = await translate_chunk(text)
    if result and user_id:
        db.add_history(user_id, text, result["translation"],
                       result["direction"])
    return result


# ==== Gửi kết quả dịch với nút lưu (callback_data đã lưu vào file) ====
async def send_translation_with_save_button(update: Update,
                                            context: ContextTypes.DEFAULT_TYPE,
                                            text: str, result: dict):
    key = short_id(text)

    # Lưu bền callback data vào file
    temp_data[key] = {"text": text, "translation": result["translation"]}
    save_temp_data(temp_data)

    keyboard = [[
        InlineKeyboardButton("💾 Lưu cụm từ này", callback_data=f"save_{key}")
    ]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    direction_map = {"zh-vi": "🇨🇳→🇻🇳", "vi-zh": "🇻🇳→🇨🇳"}
    response = f"{direction_map.get(result['direction'], '')}\n🔤 {result['translation']}\n"

    if result["pinyin"]:
        response += f"🗣️ Phiên âm: {result['pinyin']}\n"

    await context.bot.send_message(chat_id=update.message.chat.id,
                                   text=response,
                                   reply_markup=reply_markup)


# ==== Xử lý callback khi nhấn nút Lưu ====
async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data.startswith("save_"):
        key = query.data[5:]
        user_id = query.from_user.id

        info = temp_data.get(key)
        if not info:
            await query.edit_message_text("❌ Không tìm thấy cụm từ để lưu.")
            return

        db.save_phrase(user_id, info["text"], info["translation"])
        await query.edit_message_text(
            text=f"{query.message.text}\n\n✅ Đã lưu: '{info['text']}'")


# ==== Xử lý lệnh từ người dùng ====
async def handle_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    user_id = update.message.from_user.id

    if '@' in text:
        text = text.split('@')[0]

    elif text == '/mysaved':
        if not is_admin(user_id):
            await update.message.reply_text("❌ Bạn không có quyền admin.")
            return

        phrases = db.data["saved_phrases"]
        if phrases:
            msg_lines = ["📚 Tất cả cụm từ đã lưu:"]
            for user_id_key, user_phrases in phrases.items():
                msg_lines.append(f"User {user_id_key}:")
                for phrase, translation in user_phrases.items():
                    msg_lines.append(f"{phrase} → {translation}")
            msg = "\n".join(msg_lines)
            await update.message.reply_text(msg[:4000])
        else:
            await update.message.reply_text("❌ Chưa có cụm từ nào được lưu.")

    elif text == '/saved':
        phrases = db.data["saved_phrases"].get(str(user_id), {})
        if phrases:
            msg = "\n".join(f"• {k} → {v}" for k, v in phrases.items())
            await update.message.reply_text(f"📚 Cụm từ đã lưu:\n{msg[:4000]}")
        else:
            await update.message.reply_text("❌ Bạn chưa lưu cụm từ nào.")

    elif text.startswith('/save '):
        phrase = text[6:].strip()
        result = await detect_and_translate(phrase, user_id)
        if result:
            db.save_phrase(user_id, phrase, result["translation"])
            await update.message.reply_text(f"✅ Đã lưu: '{phrase}'")
        else:
            await update.message.reply_text("⚠️ Không thể dịch cụm từ này.")

    elif text.startswith('/delete '):
        phrase = text[8:].strip()
        if db.delete_phrase(user_id, phrase):
            await update.message.reply_text(f"✅ Đã xóa: '{phrase}'")
        else:
            await update.message.reply_text("❌ Không tìm thấy cụm từ này.")

    elif text.startswith('/find '):
        keyword = text[6:].strip()
        results = db.search_phrases(user_id, keyword)
        if results:
            msg = "\n".join(f"• {k} → {v}" for k, v in results.items())
            await update.message.reply_text(
                f"🔍 Kết quả tìm kiếm:\n{msg[:4000]}")
        else:
            await update.message.reply_text("❌ Không tìm thấy cụm từ.")

    elif text == '/history':
        history = db.get_user_history(user_id)
        if history:
            msg = "\n".join(
                f"{i['original']} → {i['translated']} ({i['direction']})"
                for i in history)
            await update.message.reply_text(f"📜 Lịch sử dịch:\n{msg[:4000]}")
        else:
            await update.message.reply_text("❌ Chưa có lịch sử dịch.")

    elif text.startswith('/addadmin '):
        if not is_admin(user_id):
            await update.message.reply_text("⚠️ Bạn không có quyền admin.")
            return
        new_admin = text.split()[1]
        admins.add(new_admin)
        save_admins(admins)
        await update.message.reply_text(f"✅ Đã thêm admin: {new_admin}")

    elif text.startswith('/removeadmin '):
        if not is_admin(user_id):
            await update.message.reply_text("⚠️ Bạn không có quyền admin.")
            return
        target = text.split()[1]
        if target in admins:
            admins.remove(target)
            save_admins(admins)
            await update.message.reply_text(f"✅ Đã xóa admin: {target}")
        else:
            await update.message.reply_text("❌ Không tìm thấy admin.")

    elif text.startswith('/fast'):
        if not is_admin(user_id):
            await update.message.reply_text("⚠️ Bạn không có quyền admin.")
            return

        parts = text.split()
        filter_date = None
        if len(parts) > 1:
            try:
                filter_date = datetime.strptime(parts[1], "%Y-%m-%d").date()
            except ValueError:
                await update.message.reply_text(
                    "❌ Ngày không hợp lệ. Định dạng đúng: /fast YYYY-MM-DD")
                return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lịch sử dịch"
        ws.append(["User ID", "Tên nick", "Username", "Original", "Timestamp"])

        for item in db.data["history"]:
            try:
                item_time = datetime.fromisoformat(item["timestamp"])
                if filter_date and item_time.date() != filter_date:
                    continue

                item_user_id = int(item["user_id"])
                user = await context.bot.get_chat(item_user_id)
                full_name = user.full_name or "Không rõ"
                username = f"@{user.username}" if user.username else "Không có"
            except:
                full_name = "Không lấy được"
                username = "Không lấy được"

            ws.append([
                item["user_id"], full_name, username, item["original"],
                item["timestamp"]
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[get_column_letter(
                col[0].column)].width = max_len + 2

        filename = "lich_su_dich.xlsx"
        wb.save(filename)
        await update.message.reply_document(open(filename, "rb"))
        os.remove(filename)

    elif text == '/secure':
        if not is_admin(user_id):
            await update.message.reply_text("⚠️ Bạn không có quyền admin.")
            return
        db.data["history"] = []
        db.data["saved_phrases"] = {}
        db.save_db()
        await update.message.reply_text("✅ Đã xóa toàn bộ dữ liệu.")

    else:
        await update.message.reply_text(
            "⚠️ Lệnh không hợp lệ hoặc chưa hỗ trợ.")


# ==== Xử lý tin nhắn văn bản ====
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    user_id = update.message.from_user.id

    if text.startswith('/'):
        return

    result = await detect_and_translate(text, user_id)
    if result:
        await send_translation_with_save_button(update, context, text, result)
    else:
        await update.message.reply_text("🔍 Không thể nhận diện ngôn ngữ.")


# ==== Lệnh /start ====
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🀄 越中翻译机器人 - Bot dịch Trung Việt\n\n"
        "📌 Cách dùng:\n"
        "- Gửi văn bản tiếng Việt/Trung để tự động dịch.\n"
        "- Nhấn nút 'Lưu' Để lưu cụm từ.\n"
        "- /save <cụm từ>: Lưu thủ công.\n"
        "- /saved: Xem các cụm từ đã lưu.\n"
        "- /delete <cụm từ>: Xóa cụm từ đã lưu.\n"
        "- /find <từ khóa>: Tìm cụm từ đã lưu.\n"
        "- /history: Xem lịch sử dịch.\n\n"
        "📌 使用方法:\n"
        "- 发送中越文文本自动翻译.\n"
        "- 点击'保存'按钮保存短语.\n"
        "- /saved: 查看已保存的短语.\n"
        "- /delete <短语>: 删除已保存的短语.\n"
        "- /find <关键词>: 搜索已保存的短语.\n"
        "- /history: 查看翻译历史.\n"
        "👉 @xukaxuka2k1 code free,fastandsecure👈")


async def admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id

    # Kiểm tra quyền admin
    if not is_admin(update.effective_user.id):
        return await update.message.reply_text("❌ Bạn không có quyền.")

    admin_commands = ("📜 **ADMIN MENU** 📜\n\n"
                      "🔹 /fast - ...\n"
                      "🔹 /secure - ...\n"
                      "🔹 /delete - Làm mới lại toàn bộ.\n"
                      "🔹 /addadmin - ID Làm admin.\n"
                      "🔹 /removeadmin - ID Xoá admin.")

    await update.message.reply_text(admin_commands, parse_mode="Markdown")


async def unknown(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "❓ Lệnh không hợp lệ. Gõ /start để xem lệnh.\n\n"
        "🎮 game Caro:\u2003\u2003@Game_carobot\n"
        "🎮 Nối chữ:\u2003\u2003\u2003@noi_chu_bot\n"
        "🀄 Google :\u2003\u2003@Dichngon_ngubot")


# ==== Khởi chạy Bot ====
def main():
    TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
    if not TOKEN:
        print("❌ Chưa thiết lập TELEGRAM_BOT_TOKEN")
        return

    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("admin", admin_command))
    app.add_handler(
        CommandHandler([
            "saved", "mysaved", "save", "delete", "find", "history",
            "addadmin", "removeadmin", "fast", "secure"
        ], handle_command))
    app.add_handler(CallbackQueryHandler(button_callback))
    app.add_handler(MessageHandler(filters.COMMAND, unknown))
    app.add_handler(
        MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    print("🤖 Bot đã khởi động...")
    app.run_polling()


if __name__ == '__main__':
    main()
